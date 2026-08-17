from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def deep_get(obj: dict[str, Any] | None, path: str, default: Any = None) -> Any:
    cur: Any = obj
    for part in path.split("."):
        if not isinstance(cur, dict) or part not in cur:
            return default
        cur = cur[part]
    return cur


def valid_price(value: Any) -> bool:
    return isinstance(value, (int, float)) and value >= 0


def string_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def first_valid_price(candidates: list[tuple[str, Any]]) -> tuple[Any, str]:
    for source, value in candidates:
        if valid_price(value):
            return value, source
    return None, ""


def join_dynamic_labels(raw: dict[str, Any]) -> str:
    labels: list[str] = []
    for label in raw.get("dynamic_act_labels") or []:
        for tag in label.get("sub_tags") or []:
            text = tag.get("text")
            if text:
                labels.append(str(text))
    return "；".join(dict.fromkeys(labels))


def price_display_text(activity: dict[str, Any] | None, value: Any) -> str:
    if not valid_price(value):
        return ""
    activity = activity or {}
    price_str = activity.get("activity_price_str")
    if not price_str or price_str == "-1":
        price_str = f"{value:g}"
    prefix = activity.get("activity_price_prefix") or ""
    suffix = activity.get("activity_price_suffix") or ""
    return f"{prefix}¥{price_str}{suffix}"


def product_front_price(row: dict[str, Any]) -> dict[str, Any]:
    raw = row.get("productRaw") or {}
    product_index = row.get("productIndex") or {}
    activity = deep_get(raw, "unify_price.activity_info") or {}
    price, source = first_valid_price(
        [
            ("productRaw.unify_price.activity_info.activity_price", activity.get("activity_price")),
            ("productIndex.frontDisplayPrice", product_index.get("frontDisplayPrice")),
            ("productRaw.min_price", raw.get("min_price")),
            ("productIndex.minPrice", product_index.get("minPrice")),
        ]
    )
    return {
        "value": price,
        "source": source,
        "display": price_display_text(activity, price),
        "activity": activity,
    }


def sku_front_price(sku: dict[str, Any]) -> dict[str, Any]:
    activity = deep_get(sku, "unify_price.activity_info") or {}
    price, source = first_valid_price(
        [
            ("sku.unify_price.activity_info.activity_price", activity.get("activity_price")),
            ("sku.price", sku.get("price")),
            ("sku.origin_price", sku.get("origin_price")),
        ]
    )
    return {
        "value": price,
        "source": source,
        "display": price_display_text(activity, price),
        "activity": activity,
    }


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"JSONL parse failed at line {line_no}: {exc}") from exc
    return rows


def auto_width(ws, max_width: int = 48) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells[:2000]:
            value = cell.value
            if value is not None:
                width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def style_sheet(ws, money_columns: set[str] | None = None) -> None:
    money_columns = money_columns or set()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value in money_columns:
            for body_cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for c in body_cell:
                    c.number_format = "0.00"
    auto_width(ws)


def append_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def build_workbook(raw_path: Path, summary_path: Path | None, output_path: Path) -> dict[str, Any]:
    rows = read_jsonl(raw_path)
    summary = {}
    if summary_path and summary_path.exists():
        summary = json.loads(summary_path.read_text(encoding="utf-8"))

    spu_seen: Counter[str] = Counter()
    for row in rows:
        raw = row.get("productRaw") or {}
        product_index = row.get("productIndex") or {}
        spu_id = string_value(raw.get("id") or product_index.get("spuId"))
        spu_seen[spu_id] += 1

    product_rows: list[dict[str, Any]] = []
    sku_rows: list[dict[str, Any]] = []
    category_counter: Counter[str] = Counter()
    category_unique: dict[str, set[str]] = defaultdict(set)
    price_source_counter: Counter[str] = Counter()
    stats = Counter()

    for idx, row in enumerate(rows, start=1):
        raw = row.get("productRaw") or {}
        product_index = row.get("productIndex") or {}
        category = row.get("category") or {}
        store = row.get("store") or {}
        account = row.get("account") or {}

        spu_id = string_value(raw.get("id") or product_index.get("spuId"))
        name = raw.get("name") if raw.get("name") is not None else product_index.get("name")
        category_name = category.get("displayName") or category.get("name") or ""
        category_counter[category_name] += 1
        if spu_id:
            category_unique[category_name].add(spu_id)

        front = product_front_price(row)
        activity = front["activity"]
        unify_price = raw.get("unify_price") or {}
        actual_price_info = unify_price.get("actual_price_info")
        dynamic_labels = join_dynamic_labels(raw)
        promotion_info = raw.get("promotion_info") or ""
        skus = raw.get("skus") or []

        if valid_price(front["value"]):
            stats["product_front_price_present"] += 1
            price_source_counter[front["source"]] += 1
        if actual_price_info:
            stats["actual_price_info_present"] += 1
        if promotion_info:
            stats["promotion_info_present"] += 1
        if dynamic_labels:
            stats["dynamic_labels_present"] += 1

        product_rows.append(
            {
                "序号": idx,
                "runId": row.get("runId", ""),
                "采集时间": row.get("ts", ""),
                "门店ID": store.get("storeId", ""),
                "门店名称": store.get("storeName", ""),
                "账号ID": account.get("accountId", ""),
                "Profile": account.get("profileId", ""),
                "CDP端口": account.get("cdpPort", ""),
                "类目i": category.get("i", ""),
                "类目j": category.get("j", ""),
                "一级/展示类目": category_name,
                "类目tag": category.get("tag", ""),
                "SPU_ID": spu_id,
                "商品名称": name,
                "是否重复SPU": "是" if spu_id and spu_seen[spu_id] > 1 else "否",
                "用户到手价": front["value"],
                "前端展示价文本": front["display"],
                "价格来源路径": front["source"],
                "商品min_price": raw.get("min_price", product_index.get("minPrice", "")),
                "基础展示价": unify_price.get("price", ""),
                "划线价": unify_price.get("underlined_price", activity.get("underline_price", "")),
                "价格后缀": activity.get("activity_price_suffix", ""),
                "优惠限购数": activity.get("quota_per_order", ""),
                "活动类型": activity.get("activity_type", ""),
                "活动ID": activity.get("act_id", ""),
                "活动价来源码": activity.get("activity_price_source", ""),
                "促销文案": promotion_info,
                "动态标签": dynamic_labels,
                "活动标签": raw.get("activity_tag", ""),
                "活动标签ID": raw.get("activity_tag_id", ""),
                "wmActivityType": raw.get("wmActivityType", ""),
                "月售": raw.get("month_saled_content", product_index.get("monthSaledContent", "")),
                "想买人数": raw.get("want_to_buy_content", ""),
                "好评率": raw.get("praise_rate", ""),
                "SKU数量": len(skus),
                "图片": raw.get("picture", product_index.get("picture", "")),
                "前端actual_price_info是否存在": "是" if actual_price_info else "否",
                "source": row.get("source", ""),
            }
        )

        for sku_i, sku in enumerate(skus, start=1):
            sku_front = sku_front_price(sku)
            sku_activity = sku_front["activity"]
            if valid_price(sku_front["value"]):
                stats["sku_front_price_present"] += 1
            sku_rows.append(
                {
                    "序号": len(sku_rows) + 1,
                    "商品序号": idx,
                    "门店名称": store.get("storeName", ""),
                    "一级/展示类目": category_name,
                    "SPU_ID": spu_id,
                    "商品名称": name,
                    "SKU序号": sku_i,
                    "SKU_ID": string_value(sku.get("id")),
                    "规格": sku.get("spec", ""),
                    "条码": sku.get("upccode", ""),
                    "SKU用户到手价": sku_front["value"],
                    "SKU前端展示价文本": sku_front["display"],
                    "SKU价格来源路径": sku_front["source"],
                    "SKU单件价": sku.get("price", ""),
                    "SKU原价": sku.get("origin_price", ""),
                    "SKU基础展示价": deep_get(sku, "unify_price.price", ""),
                    "SKU划线价": deep_get(sku, "unify_price.underlined_price", sku_activity.get("underline_price", "")),
                    "起购数量": sku.get("min_order_count", ""),
                    "限购数量": sku_activity.get("quota_per_order", ""),
                    "SKU促销文案": sku.get("promotion_info", ""),
                    "SKU活动类型": sku_activity.get("activity_type", ""),
                    "SKU活动ID": sku_activity.get("act_id", ""),
                    "库存": sku.get("stock", ""),
                    "真实库存": sku.get("real_stock", ""),
                    "活动库存": sku.get("activity_stock", ""),
                    "状态": sku.get("status", ""),
                }
            )

    category_rows = []
    for category_name, count in category_counter.most_common():
        category_rows.append(
            {
                "一级/展示类目": category_name,
                "采集行数": count,
                "唯一SPU数": len(category_unique[category_name]),
            }
        )

    overview_rows = [
        {"指标": "导出时间", "值": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"指标": "原始JSONL", "值": str(raw_path)},
        {"指标": "采集runId", "值": summary.get("runId", rows[0].get("runId", "") if rows else "")},
        {"指标": "门店名称", "值": summary.get("storeName", rows[0].get("store", {}).get("storeName", "") if rows else "")},
        {"指标": "采集状态", "值": summary.get("status", "")},
        {"指标": "状态说明", "值": "failed 是人工暂停 stop file 导致，不代表原始数据损坏" if summary.get("error", "").startswith("stop file exists") else summary.get("error", "")},
        {"指标": "原始行数", "值": len(rows)},
        {"指标": "唯一SPU数", "值": len(spu_seen)},
        {"指标": "SKU明细行数", "值": len(sku_rows)},
        {"指标": "请求数", "值": summary.get("requestCount", "")},
        {"指标": "风险事件数", "值": summary.get("riskCount", "")},
        {"指标": "商品用户到手价覆盖", "值": f"{stats['product_front_price_present']}/{len(rows)}"},
        {"指标": "SKU用户到手价覆盖", "值": f"{stats['sku_front_price_present']}/{len(sku_rows)}"},
        {"指标": "actual_price_info覆盖", "值": f"{stats['actual_price_info_present']}/{len(rows)}"},
        {"指标": "促销文案覆盖", "值": f"{stats['promotion_info_present']}/{len(rows)}"},
        {"指标": "动态标签覆盖", "值": f"{stats['dynamic_labels_present']}/{len(rows)}"},
        {"指标": "用户到手价主来源", "值": "; ".join(f"{k}: {v}" for k, v in price_source_counter.items())},
    ]

    field_rows = [
        {"字段": "用户到手价", "说明": "当前最可信口径为 productRaw.unify_price.activity_info.activity_price；本批次 2798 行均有覆盖。"},
        {"字段": "前端展示价文本", "说明": "由 activity_price_str + activity_price_suffix 组装，例如 ¥0.01起，用于还原前端价签观感。"},
        {"字段": "商品min_price", "说明": "保留原始 min_price 供排查，但不能直接作为用户到手价；本批次已观察到两者大量不一致。"},
        {"字段": "actual_price_info", "说明": "仅少量商品存在，不能作为主口径；可用于后续校验复杂满减/会员场景。"},
        {"字段": "促销文案/动态标签", "说明": "用于解释低价来源，例如限购、任选、折扣标签；不做清洗，最大程度保留前端原始信息。"},
        {"字段": "SKU明细", "说明": "SKU级价格可能受起购数量影响，已同时保留 SKU活动价、单件价、起购数量、限购数量。"},
    ]

    wb = Workbook()
    ws_product = wb.active
    ws_product.title = "商品SPU"
    product_headers = [
        "序号",
        "runId",
        "采集时间",
        "门店ID",
        "门店名称",
        "账号ID",
        "Profile",
        "CDP端口",
        "类目i",
        "类目j",
        "一级/展示类目",
        "类目tag",
        "SPU_ID",
        "商品名称",
        "是否重复SPU",
        "用户到手价",
        "前端展示价文本",
        "价格来源路径",
        "商品min_price",
        "基础展示价",
        "划线价",
        "价格后缀",
        "优惠限购数",
        "活动类型",
        "活动ID",
        "活动价来源码",
        "促销文案",
        "动态标签",
        "活动标签",
        "活动标签ID",
        "wmActivityType",
        "月售",
        "想买人数",
        "好评率",
        "SKU数量",
        "图片",
        "前端actual_price_info是否存在",
        "source",
    ]
    append_rows(ws_product, product_headers, product_rows)
    style_sheet(ws_product, {"用户到手价", "商品min_price", "基础展示价", "划线价"})

    ws_sku = wb.create_sheet("SKU明细")
    sku_headers = [
        "序号",
        "商品序号",
        "门店名称",
        "一级/展示类目",
        "SPU_ID",
        "商品名称",
        "SKU序号",
        "SKU_ID",
        "规格",
        "条码",
        "SKU用户到手价",
        "SKU前端展示价文本",
        "SKU价格来源路径",
        "SKU单件价",
        "SKU原价",
        "SKU基础展示价",
        "SKU划线价",
        "起购数量",
        "限购数量",
        "SKU促销文案",
        "SKU活动类型",
        "SKU活动ID",
        "库存",
        "真实库存",
        "活动库存",
        "状态",
    ]
    append_rows(ws_sku, sku_headers, sku_rows)
    style_sheet(ws_sku, {"SKU用户到手价", "SKU单件价", "SKU原价", "SKU基础展示价", "SKU划线价"})

    ws_category = wb.create_sheet("类目汇总")
    append_rows(ws_category, ["一级/展示类目", "采集行数", "唯一SPU数"], category_rows)
    style_sheet(ws_category)

    ws_overview = wb.create_sheet("采集概况")
    append_rows(ws_overview, ["指标", "值"], overview_rows)
    style_sheet(ws_overview)

    ws_fields = wb.create_sheet("字段说明")
    append_rows(ws_fields, ["字段", "说明"], field_rows)
    style_sheet(ws_fields)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    duplicate_spu_count = sum(1 for count in spu_seen.values() if count > 1)
    completeness_status = "pass"
    if len(rows) == 0 or stats["product_front_price_present"] == 0:
        completeness_status = "fail"
    elif stats["product_front_price_present"] < len(rows) or stats["sku_front_price_present"] < len(sku_rows):
        completeness_status = "warn"

    verified = load_workbook(output_path, read_only=True, data_only=True)
    verify_rows = {name: verified[name].max_row for name in verified.sheetnames}
    verified.close()
    return {
        "output": str(output_path),
        "sheets": verify_rows,
        "rawRows": len(rows),
        "uniqueSpu": len(spu_seen),
        "skuRows": len(sku_rows),
        "productFrontPriceCoverage": f"{stats['product_front_price_present']}/{len(rows)}",
        "skuFrontPriceCoverage": f"{stats['sku_front_price_present']}/{len(sku_rows)}",
        "quality": {
            "storeId": summary.get("storeId"),
            "workerId": (rows[0].get("worker") or {}).get("workerId") if rows else None,
            "accountId": summary.get("accountId"),
            "profileId": summary.get("profileId"),
            "rawRows": len(rows),
            "uniqueSpuCount": len(spu_seen),
            "skuRows": len(sku_rows),
            "frontDisplayPricePresent": stats["product_front_price_present"],
            "skuFrontDisplayPricePresent": stats["sku_front_price_present"],
            "actualPriceInfoPresent": stats["actual_price_info_present"],
            "promotionInfoPresent": stats["promotion_info_present"],
            "dynamicLabelPresent": stats["dynamic_labels_present"],
            "duplicateSpuCount": duplicate_spu_count,
            "completenessStatus": completeness_status,
            "metadata": {
                "runId": summary.get("runId"),
                "storeName": summary.get("storeName"),
                "excelPath": str(output_path),
                "rawPath": str(raw_path),
                "priceSourceCounter": dict(price_source_counter),
                "summaryStatus": summary.get("status"),
                "summaryError": summary.get("error")
            },
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Meituan raw JSONL to an Excel workbook with front-display price fields.")
    parser.add_argument("--raw", required=True, type=Path)
    parser.add_argument("--summary", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quality-output", type=Path)
    args = parser.parse_args()
    result = build_workbook(args.raw, args.summary, args.output)
    if args.quality_output:
        args.quality_output.parent.mkdir(parents=True, exist_ok=True)
        args.quality_output.write_text(json.dumps(result["quality"], ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

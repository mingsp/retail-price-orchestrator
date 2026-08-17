from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_MIN_PRICE_COVERAGE = 0.995


def deep_get(obj: dict[str, Any] | None, path: str, default: Any = None) -> Any:
    cur: Any = obj
    for part in path.split("."):
        if not isinstance(cur, dict) or part not in cur:
            return default
        cur = cur[part]
    return cur


def valid_price(value: Any) -> bool:
    return isinstance(value, (int, float)) and value >= 0


def front_price(raw: dict[str, Any], fallback: Any = None) -> tuple[Any, str]:
    activity = deep_get(raw, "unify_price.activity_info") or {}
    value = activity.get("activity_price")
    if not valid_price(value):
        value = raw.get("min_price")
    if not valid_price(value):
        value = fallback
    price_str = activity.get("activity_price_str")
    if not price_str or price_str == "-1":
        price_str = f"{value:g}" if valid_price(value) else ""
    prefix = activity.get("activity_price_prefix") or ""
    suffix = activity.get("activity_price_suffix") or ""
    display = f"{prefix}¥{price_str}{suffix}" if valid_price(value) else ""
    return value, display


def sku_front_price(sku: dict[str, Any]) -> tuple[Any, str]:
    activity = deep_get(sku, "unify_price.activity_info") or {}
    value = activity.get("activity_price")
    if not valid_price(value):
        value = sku.get("price")
    price_str = activity.get("activity_price_str")
    if not price_str or price_str == "-1":
        price_str = f"{value:g}" if valid_price(value) else ""
    prefix = activity.get("activity_price_prefix") or ""
    suffix = activity.get("activity_price_suffix") or ""
    display = f"{prefix}¥{price_str}{suffix}" if valid_price(value) else ""
    return value, display


def join_dynamic_labels(raw: dict[str, Any]) -> str:
    labels: list[str] = []
    for label in raw.get("dynamic_act_labels") or []:
        for tag in label.get("sub_tags") or []:
            text = tag.get("text")
            if text:
                labels.append(str(text))
    return "；".join(dict.fromkeys(labels))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def read_quality_manifest(path: Path | None) -> dict[str, Any] | None:
    if not path:
        return None
    body = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(body, dict) and isinstance(body.get("quality"), dict):
        return body["quality"]
    if isinstance(body, dict):
        return body
    raise ValueError("quality_manifest_invalid")


def validate_quality_gate(
    *,
    quality: dict[str, Any] | None,
    price_present: int,
    product_count: int,
    min_price_coverage: float,
) -> None:
    if quality:
        status = quality.get("completenessStatus")
        if status == "fail":
            raise ValueError("quality_gate_failed: completenessStatus=fail")
    if product_count <= 0:
        raise ValueError("quality_gate_failed: no_products")
    coverage = price_present / product_count
    if coverage < min_price_coverage:
        raise ValueError(f"quality_gate_failed: priceCoverage={price_present}/{product_count}")


def append_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def style_sheet(ws, money_columns: set[str] | None = None) -> None:
    money_columns = money_columns or set()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value in money_columns:
            for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for c in col:
                    c.number_format = "0.00"
    for cells in ws.columns:
        letter = get_column_letter(cells[0].column)
        width = 10
        for cell in cells[:2000]:
            if cell.value is not None:
                width = max(width, min(46, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def build_business_workbook(
    raw_path: Path,
    summary_path: Path | None,
    output_path: Path,
    quality_path: Path | None = None,
    min_price_coverage: float = DEFAULT_MIN_PRICE_COVERAGE,
) -> dict[str, Any]:
    raw_rows = read_jsonl(raw_path)
    summary = json.loads(summary_path.read_text(encoding="utf-8")) if summary_path and summary_path.exists() else {}
    quality = read_quality_manifest(quality_path)

    products: list[dict[str, Any]] = []
    skus: list[dict[str, Any]] = []
    category_prices: dict[str, list[float]] = defaultdict(list)
    category_spus: dict[str, set[str]] = defaultdict(set)
    category_counts: Counter[str] = Counter()
    price_present = 0
    sku_price_present = 0

    for row in raw_rows:
        raw = row.get("productRaw") or {}
        index = row.get("productIndex") or {}
        category = row.get("category") or {}
        store = row.get("store") or {}
        category_name = category.get("displayName") or category.get("name") or ""
        name = raw.get("name") if raw.get("name") is not None else index.get("name")
        spu_id = str(raw.get("id") or index.get("spuId") or "")
        value, display = front_price(raw, index.get("minPrice"))
        activity = deep_get(raw, "unify_price.activity_info") or {}
        unify = raw.get("unify_price") or {}
        promotion = raw.get("promotion_info") or ""
        labels = join_dynamic_labels(raw)
        sku_list = raw.get("skus") or []

        if valid_price(value):
            price_present += 1
            category_prices[category_name].append(float(value))
        category_spus[category_name].add(spu_id)
        category_counts[category_name] += 1

        products.append(
            {
                "门店": store.get("storeName") or summary.get("storeName", ""),
                "类目": category_name,
                "商品名称": name,
                "用户到手价": value,
                "前端展示价": display,
                "基础展示价": unify.get("price", ""),
                "划线价": unify.get("underlined_price", activity.get("underline_price", "")),
                "优惠/活动": promotion,
                "前端标签": labels,
                "限购数量": activity.get("quota_per_order", ""),
                "月售": raw.get("month_saled_content", index.get("monthSaledContent", "")),
                "想买/热度": raw.get("want_to_buy_content", ""),
                "好评率": raw.get("praise_rate", ""),
                "规格数": len(sku_list),
                "图片": raw.get("picture", index.get("picture", "")),
            }
        )

        for sku in sku_list:
            sku_value, sku_display = sku_front_price(sku)
            sku_activity = deep_get(sku, "unify_price.activity_info") or {}
            if valid_price(sku_value):
                sku_price_present += 1
            skus.append(
                {
                    "门店": store.get("storeName") or summary.get("storeName", ""),
                    "类目": category_name,
                    "商品名称": name,
                    "规格": sku.get("spec", ""),
                    "条码": sku.get("upccode", ""),
                    "SKU到手价": sku_value,
                    "SKU前端展示价": sku_display,
                    "SKU单件价": sku.get("price", ""),
                    "SKU原价": sku.get("origin_price", ""),
                    "起购数量": sku.get("min_order_count", ""),
                    "限购数量": sku_activity.get("quota_per_order", ""),
                    "SKU优惠/活动": sku.get("promotion_info", ""),
                    "库存": sku.get("stock", ""),
                    "图片": sku.get("picture", raw.get("picture", "")),
                }
            )

    category_rows = []
    for category_name, count in category_counts.most_common():
        prices = category_prices.get(category_name, [])
        category_rows.append(
            {
                "类目": category_name,
                "商品条数": count,
                "去重商品数": len(category_spus[category_name]),
                "平均用户到手价": round(mean(prices), 2) if prices else "",
                "最低用户到手价": min(prices) if prices else "",
                "最高用户到手价": max(prices) if prices else "",
            }
        )

    note_rows = [
        {"项目": "门店", "内容": summary.get("storeName", products[0]["门店"] if products else "")},
        {"项目": "导出时间", "内容": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"项目": "商品主表行数", "内容": len(products)},
        {"项目": "SKU明细行数", "内容": len(skus)},
        {"项目": "用户到手价覆盖", "内容": f"{price_present}/{len(products)}"},
        {"项目": "SKU到手价覆盖", "内容": f"{sku_price_present}/{len(skus)}"},
        {"项目": "质量闸", "内容": f"用户到手价覆盖率阈值 {min_price_coverage:.1%}；quality manifest 状态为 fail 时阻断导出。"},
        {"项目": "说明", "内容": "本表为同事使用版，已移除 runId、账号、Profile、CDP、接口路径等采集追踪字段。商品名称保持采集原文，不做清洗改写。"},
    ]

    wb = Workbook()
    product_sheet = wb.active
    product_sheet.title = "商品清单"
    append_rows(
        product_sheet,
        [
            "门店",
            "类目",
            "商品名称",
            "用户到手价",
            "前端展示价",
            "基础展示价",
            "划线价",
            "优惠/活动",
            "前端标签",
            "限购数量",
            "月售",
            "想买/热度",
            "好评率",
            "规格数",
            "图片",
        ],
        products,
    )
    style_sheet(product_sheet, {"用户到手价", "基础展示价", "划线价"})

    sku_sheet = wb.create_sheet("SKU规格明细")
    append_rows(
        sku_sheet,
        [
            "门店",
            "类目",
            "商品名称",
            "规格",
            "条码",
            "SKU到手价",
            "SKU前端展示价",
            "SKU单件价",
            "SKU原价",
            "起购数量",
            "限购数量",
            "SKU优惠/活动",
            "库存",
            "图片",
        ],
        skus,
    )
    style_sheet(sku_sheet, {"SKU到手价", "SKU单件价", "SKU原价"})

    category_sheet = wb.create_sheet("类目汇总")
    append_rows(
        category_sheet,
        ["类目", "商品条数", "去重商品数", "平均用户到手价", "最低用户到手价", "最高用户到手价"],
        category_rows,
    )
    style_sheet(category_sheet, {"平均用户到手价", "最低用户到手价", "最高用户到手价"})

    note_sheet = wb.create_sheet("说明")
    append_rows(note_sheet, ["项目", "内容"], note_rows)
    style_sheet(note_sheet)

    validate_quality_gate(
        quality=quality,
        price_present=price_present,
        product_count=len(products),
        min_price_coverage=min_price_coverage,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    verified = load_workbook(output_path, read_only=True, data_only=True)
    sheets = {name: verified[name].max_row for name in verified.sheetnames}
    verified.close()
    return {
        "output": str(output_path),
        "sheets": sheets,
        "products": len(products),
        "skus": len(skus),
        "priceCoverage": f"{price_present}/{len(products)}",
        "skuPriceCoverage": f"{sku_price_present}/{len(skus)}",
        "qualityGate": "pass",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Export colleague-facing Meituan price workbook.")
    parser.add_argument("--raw", required=True, type=Path)
    parser.add_argument("--summary", type=Path)
    parser.add_argument("--quality", type=Path, help="Optional quality manifest; completenessStatus=fail blocks export.")
    parser.add_argument("--min-price-coverage", type=float, default=DEFAULT_MIN_PRICE_COVERAGE)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    print(json.dumps(
        build_business_workbook(args.raw, args.summary, args.output, args.quality, args.min_price_coverage),
        ensure_ascii=False,
        indent=2
    ))


if __name__ == "__main__":
    main()
